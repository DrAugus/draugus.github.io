# 要使用Python解析Excel文件，需要使用Python的第三方库，比如`xlrd`和`openpyxl`。

# `openpyxl`库适用于读写Excel 2007及以后版本的.xlsx格式文件。下面是读取Excel文件的示例代码：

from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from collections import OrderedDict

current_path = os.path.dirname(__file__)


def fps_format():

    filename = "script/test.xlsx"

    outputfile = "script/output.xlsx"

    # 打开Excel文件
    wb = load_workbook(filename)

    # 获取所有sheet的名称
    print(wb.sheetnames)

    # 选择要读取的sheet
    sheet1 = wb["Result 1"]

    # 获取行数和列数
    nrows = sheet1.max_row
    ncols = sheet1.max_column

    # 遍历所有行和列
    # for row in sheet1.iter_rows(values_only=True):
    #     for cell_value in row:
    #         print(cell_value)

    # /home/test_fps/gene/fps15/fps15_7204997026155744572.4822.flv.far
    # /home/test_fps/gene/fps/7204997026155744572.4822.flv.far
    # /home/test_fps/gene/fps10/fps10_7204997026155744572.4822.flv.far
    # /home/test_fps/gene/fps1/fps1_7204997026155744572.4822.flv.far
    # /home/test_fps/gene/fps5/gene_in.sh_fps5_7204997026155744572.4822.flv.far

    def classify_value(str, num, res_value):
        match_arr = ["fps", "fps15", "fps10", "fps5", "fps1"]
        for idx, ma in enumerate(match_arr):
            ss = f"/home/test_fps/gene/{ma}/"
            if str.startswith(ss) and num > 0:
                res_value[idx][1] = 1

    # 读取Excel文件到DataFrame对象
    df = pd.read_excel(filename)

    # 将DataFrame对象转换为字典
    # result = dict(zip(df['site_asset_id'], df['meta_title']))

    # site_asset_id meta_title video_duration
    # 将A列作为key，B列和C列组成一个tuple作为value，放入字典中
    data_dict = {}
    for index, row in df.iterrows():
        key = row["site_asset_id"]
        value = (row["meta_title"], row["video_duration"])
        if key not in data_dict:
            data_dict[key] = []
        data_dict[key].append(value)

    print(data_dict)

    res_dict = {}
    for key, value in data_dict.items():
        res_key = key
        res_value = [
            ["fps", 0],
            ["fps15", 0],
            ["fps10", 0],
            ["fps5", 0],
            ["fps1", 0],
        ]

        # print(f"Key: {key}")
        for v in value:
            classify_value(v[0], v[1], res_value)
            # print(f"Value - B: {v[0]}, C: {v[1]}")

        if res_key not in res_dict:
            res_dict[res_key] = []
        res_dict[res_key] = res_value

    print(res_dict)

    num_keys = len(data_dict.keys())
    print(num_keys)

    # 创建一个空DataFrame
    res_df = pd.DataFrame(
        columns=["key"]
        + list(
            OrderedDict.fromkeys(
                [pair[0] for pairs in res_dict.values() for pair in pairs]
            )
        )
    )

    # 遍历字典，将数据填入DataFrame中
    dataframes = []

    for k, v in res_dict.items():
        values = {pair[0]: pair[1] for pair in v}
        values["key"] = k
        dataframes.append(pd.DataFrame(values, index=[0]))

    res_df = pd.concat(dataframes, ignore_index=True)

    # 将DataFrame写入Excel文件
    res_df.to_excel(outputfile, index=False)

    tttt = {
        "fps1.7096759473913548047.2704.fbl.flv": [
            ["fps", 1],
            ["fps1", 0],
            ["fps5", 1],
            ["fps10", 1],
            ["fps15", 1],
        ],
        "fps5.7096759473913548047.2704.fbl.flv": [
            ["fps", 1],
            ["fps1", 1],
            ["fps5", 0],
            ["fps10", 1],
            ["fps15", 1],
        ],
        "fps15.7110599057562733858.21359.fbl.flv": [
            ["fps", 1],
            ["fps1", 1],
            ["fps5", 1],
            ["fps10", 1],
            ["fps15", 1],
        ],
    }


# "January 9, 2007" to "2007/01/09"
def format_date(date_str):

    if not isinstance(date_str, str) or not len(date_str):
        return date_str

    # 解析原始日期字符串
    date_obj = datetime.strptime(date_str, "%B %d, %Y")

    # 格式化为 "YYYY/MM/DD" 格式
    formatted_date = date_obj.strftime("%Y/%m/%d")
    return formatted_date


def read_excel():

    # 读取Excel文件
    filename = current_path + "/test.xlsx"
    sheet_name = "Sheet1"  # 替换为你的工作表名称

    df = pd.read_excel(filename, sheet_name=sheet_name)

    df["formatted_announced"] = df["announced"].apply(format_date)

    # 获取前三列数据并转换为字典数组
    # 列名映射，确保它们对应Excel表格中的前三列
    columns = ["generation", "model", "formatted_announced"]
    result = df[columns].to_dict("records")

    # 输出结果
    print(result)
    return result


def modify_res(rs):
    res = []
    gen = 0
    models = []
    dic = {"generation": gen, "model": models, "formatted_announced": ""}
    for item in rs:
        if item.generation != "nan":
            gen = item.generation
            dic["model"] = []
            dic["model"].append(item.model)
        else:
            dic["model"].append(item.model)


read_excel()
